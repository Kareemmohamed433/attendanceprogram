from flask import Flask, render_template, Response, redirect, url_for, request, session, flash
from flask_session import Session
import cv2
import face_recognition
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from functools import wraps

app = Flask(__name__)
app.secret_key = "your_secret_key_here"  # استبدلها بقيمة آمنة

# إعداد Flask-Session لتخزين بيانات الجلسة على الخادم باستخدام نظام الملفات
app.config['SESSION_TYPE'] = 'filesystem'
Session(app)

# بيانات المستخدمين (يمكن تعديلها لاحقاً أو استبدالها بقاعدة بيانات)
users = {
    "admin": {"password": "admin123", "role": "admin"},
    "user1": {"password": "password1", "role": "user"}
}

# ديكوريتور لحماية الصفحات التي تتطلب تسجيل الدخول
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash("يرجى تسجيل الدخول للوصول إلى هذه الصفحة", "warning")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# مسار تسجيل المستخدمين الجدد (متاح فقط للمسؤول)
@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    if session.get('user') != 'admin':
        flash("ليس لديك صلاحية للوصول إلى صفحة التسجيل", "danger")
        return redirect(url_for('home'))
    if request.method == 'POST':
        new_username = request.form.get('username')
        new_password = request.form.get('password')
        new_role = request.form.get('role', 'user')
        if new_username in users:
            flash("اسم المستخدم موجود بالفعل", "warning")
        else:
            users[new_username] = {'password': new_password, 'role': new_role}
            flash("تم إضافة المستخدم بنجاح", "success")
            return redirect(url_for('home'))
    return render_template('register.html')

# مسار إدارة المستخدمين (متاح فقط للمسؤول)
@app.route('/manage_users')
@login_required
def manage_users():
    if session.get('user') != 'admin':
        flash("ليس لديك صلاحية الوصول إلى هذه الصفحة", "danger")
        return redirect(url_for('home'))
    return render_template('manage_users.html', users=users)

# مسار حذف مستخدم (باستثناء المستخدم "admin")
@app.route('/delete_user/<username>', methods=['POST'])
@login_required
def delete_user(username):
    if session.get('user') != 'admin':
        flash("ليس لديك صلاحية تنفيذ هذه العملية", "danger")
        return redirect(url_for('home'))
    if username == 'admin':
        flash("لا يمكن حذف المستخدم الإداري", "danger")
    elif username in users:
        del users[username]
        flash("تم حذف المستخدم بنجاح", "success")
    else:
        flash("المستخدم غير موجود", "danger")
    return redirect(url_for('manage_users'))

# ----------------------------------
# إعداد التعرف على الوجوه والحضور
# ----------------------------------

face_encodings_list = []
face_names = []

photo_dir = "photo/"
if not os.path.exists(photo_dir):
    print(f"⚠️ خطأ: المجلد '{photo_dir}' غير موجود.")
    exit()

for file_name in os.listdir(photo_dir):
    if file_name.lower().endswith((".jpg", ".jpeg", ".png")):
        image_path = os.path.join(photo_dir, file_name)
        image = face_recognition.load_image_file(image_path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            face_encodings_list.append(encodings[0])
            face_names.append(os.path.splitext(file_name)[0])
        else:
            print(f"⚠️ تحذير: لم يتم العثور على وجه في {file_name}")

if not face_encodings_list:
    print("⚠️ خطأ: لم يتم تحميل أي وجوه، تأكد من وضع صور صالحة في المجلد.")
    exit()

all_students = face_names.copy()
attended_students = []
last_attendance_day = {student: None for student in all_students}
monthly_attendance = {student: 0 for student in all_students}
payment_status = {student: False for student in all_students}
payment_method = {student: None for student in all_students}

capturing = False  # حالة تسجيل الحضور
fee = 100          # قيمة الرسوم لكل طالب

# بدء التقاط الفيديو من الكاميرا
video_cap = cv2.VideoCapture(0)
video_cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
video_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

def gen_frames():
    """توليد إطارات الفيديو لمعالجتها وبثها عبر الويب."""
    global attended_students, capturing, last_attendance_day
    while True:
        ret, frame = video_cap.read()
        if not ret:
            break

        if capturing:
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            face_locations = face_recognition.face_locations(rgb)
            face_encodings = face_recognition.face_encodings(rgb, face_locations)
            face_names_detected = []
            today = datetime.now().strftime("%Y-%m-%d")

            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(face_encodings_list, face_encoding, tolerance=0.5)
                face_distances = face_recognition.face_distance(face_encodings_list, face_encoding)
                if any(matches):
                    best_match = np.argmin(face_distances)
                    name = face_names[best_match]
                else:
                    name = "Unknown"
                face_names_detected.append(name)
                if name != "Unknown":
                    if last_attendance_day.get(name) != today:
                        attended_students.append(name)
                        last_attendance_day[name] = today

            for (top, right, bottom, left), name in zip(face_locations, face_names_detected):
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.putText(frame, name, (left, top - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
        else:
            cv2.putText(frame, "Attendance Not Started", (50, 50),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

        ret, buffer = cv2.imencode('.jpg', frame)
        frame_bytes = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')

# ----------------------------------
# مسارات التطبيق الأخرى
# ----------------------------------

# الصفحة الرئيسية: إعادة التوجيه إلى تسجيل الدخول إذا لم يكن المستخدم مسجلاً
@app.route('/')
def home():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', capturing=capturing, fee=fee)

# صفحة تسجيل الدخول
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = users.get(username)
        if user and user['password'] == password:
            session['user'] = username
            flash("تم تسجيل الدخول بنجاح", "success")
            return redirect(url_for('home'))
        else:
            flash("اسم المستخدم أو كلمة المرور غير صحيحة", "danger")
    return render_template('login.html')

# تسجيل الخروج
@app.route('/logout')
def logout():
    session.pop('user', None)
    flash("تم تسجيل الخروج", "info")
    return redirect(url_for('login'))

@app.route('/video_feed')
@login_required
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/start')
@login_required
def start_attendance():
    global capturing
    capturing = True
    return redirect(url_for('home'))

@app.route('/end')
@login_required
def end_attendance():
    global capturing, attended_students, monthly_attendance
    capturing = False
    today = datetime.now().strftime("%Y-%m-%d")
    
    for student in set(attended_students):
        monthly_attendance[student] += 1

    excel_file = "Attendance.xlsx"
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Student")
        for i, student in enumerate(all_students, start=2):
            ws.cell(row=i, column=1, value=student)
    
    col_attendance = None
    for c in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == today:
            col_attendance = c
            break
    if not col_attendance:
        col_attendance = ws.max_column + 1
        ws.cell(row=1, column=col_attendance, value=today)
    
    for i, student in enumerate(all_students, start=2):
        ws.cell(row=i, column=col_attendance, value=True if student in set(attended_students) else False)
    
    payment_col = None
    for c in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "Payment Status":
            payment_col = c
            break
    if not payment_col:
        payment_col = ws.max_column + 1
        ws.cell(row=1, column=payment_col, value="Payment Status")
    
    for i, student in enumerate(all_students, start=2):
        status = "Paid" if payment_status.get(student, False) else "Not Paid"
        ws.cell(row=i, column=payment_col, value=status)
    
    method_col = None
    for c in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "Payment Method":
            method_col = c
            break
    if not method_col:
        method_col = ws.max_column + 1
        ws.cell(row=1, column=method_col, value="Payment Method")
    
    for i, student in enumerate(all_students, start=2):
        method = payment_method.get(student, "") if payment_status.get(student, False) else ""
        ws.cell(row=i, column=method_col, value=method)
    
    total_col = None
    for c in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "اجمالي":
            total_col = c
            break
    if not total_col:
        total_col = ws.max_column + 1
        ws.cell(row=1, column=total_col, value="اجمالي")
    
    for i, student in enumerate(all_students, start=2):
        total_value = fee if payment_status.get(student, False) else 0
        ws.cell(row=i, column=total_col, value=total_value)
    
    wb.save(excel_file)
    attended_students = []
    return redirect(url_for('home'))

@app.route('/mark_paid/<student>', methods=['GET', 'POST'])
@login_required
def mark_paid(student):
    global payment_status, payment_method
    if request.method == 'POST':
        method = request.form.get('payment_method')
        if student in payment_status:
            payment_status[student] = True
            payment_method[student] = method
        return redirect(url_for('attendance_status'))
    return render_template('mark_paid.html', student=student)

@app.route('/attendance_status')
@login_required
def attendance_status():
    monthly_data = []
    for student in all_students:
        count = monthly_attendance.get(student, 0)
        message = ""
        if count < 6:
            message = f"لم يكمل الشهر ({count}/6)"
        pay_status = "Paid" if payment_status.get(student, False) else "Not Paid"
        method = payment_method.get(student, "") if payment_status.get(student, False) else ""
        monthly_data.append({"name": student, "count": count, "message": message, "paid": pay_status, "method": method})
    return render_template('attendance.html', monthly_data=monthly_data, fee=fee)

@app.route('/total_payment')
@login_required
def total_payment():
    global fee
    paid_count = sum(1 for s in payment_status if payment_status[s])
    total = paid_count * fee
    return render_template('total_payment.html', paid_count=paid_count, fee=fee, total=total)

@app.route('/set_fee', methods=['GET', 'POST'])
@login_required
def set_fee():
    global fee
    if request.method == 'POST':
        try:
            fee = float(request.form.get('fee'))
        except (ValueError, TypeError):
            fee = 100
        return redirect(url_for('home'))
    return render_template('set_fee.html', fee=fee)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)



