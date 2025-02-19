{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": 1,
   "id": "16e75bc1",
   "metadata": {
    "scrolled": true
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "๐ท ุงููุงููุฑุง ุชุนูู... ุงุถุบุท 'q' ููุฅุบูุงู\n",
      "ุชู ุชุญููู ุงูุฃุณูุงุก ุงูุชุงููุฉ: ['face_down', 'face_front', 'face_left', 'face_right', 'face_up']\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.47403225 0.47397465 0.47887733 0.47930566 0.45586706]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.40933962 0.42847192 0.40006674 0.43470373 0.40701798]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.356573   0.36259352 0.33451405 0.37241694 0.37777423]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.40463002 0.41790161 0.36507556 0.41708691 0.41503048]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.39687243 0.43955492 0.40378566 0.41402875 0.41581826]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.47641812 0.45748055 0.4783522  0.4784571  0.47409178]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.42485795 0.46336804 0.46188895 0.45058214 0.45422008]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.41343592 0.42955727 0.41743048 0.42613336 0.41786397]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.43793972 0.43917321 0.41915006 0.42485792 0.41621797]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.43640946 0.42647529 0.41373419 0.42108525 0.39486386]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.44525152 0.42447041 0.41833129 0.42339321 0.41251715]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.42827375 0.43319469 0.40939395 0.42111217 0.40363284]\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.43475801 0.42474686 0.41147902 0.42045895 0.40215339]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.44440924 0.42874272 0.42161161 0.42481868 0.41354007]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.45490873 0.42564278 0.42097444 0.42501569 0.41671242]\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.56455349 0.48285205 0.50075227 0.49264499 0.52436651]\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.5035921  0.45199477 0.46275382 0.44030816 0.43550158]\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.52205181 0.43835179 0.44388418 0.44183781 0.45497035]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.4551323  0.38494386 0.38543064 0.39349649 0.41544009]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.51006733 0.43164533 0.44519455 0.43914828 0.44381206]\n",
      "โ ุชู ุงูุชุดุงู 0 ูุฌู(ูุฌูู)\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.49406669 0.40634652 0.42467452 0.41656708 0.43893903]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.50085707 0.4235097  0.43688968 0.42848284 0.42675776]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.49903588 0.4019664  0.41896137 0.41795689 0.4018736 ]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.53101999 0.45920586 0.47567208 0.45934557 0.48592176]\n",
      "๐ด ุฅุบูุงู ุงููุงููุฑุง.\n",
      "โ ุชู ุฅุบูุงู ุงููุงููุฑุง ูุฅููุงู ุงูุจุฑูุงูุฌ ุจูุฌุงุญ.\n"
     ]
    }
   ],
   "source": [
    "import cv2\n",
    "import face_recognition\n",
    "import openpyxl\n",
    "from datetime import datetime\n",
    "import numpy as np\n",
    "import os\n",
    "\n",
    "# ุจุฏุก ุงูุชูุงุท ุงูููุฏูู ูู ุงููุงููุฑุง\n",
    "video_cap = cv2.VideoCapture(0)\n",
    "\n",
    "# ุงูุชุญูู ูู ุฃู ุงููุงููุฑุง ุชุนูู\n",
    "if not video_cap.isOpened():\n",
    "    print(\"โ๏ธ ุฎุทุฃ: ุชุนุฐุฑ ูุชุญ ุงููุงููุฑุง.\")\n",
    "    exit()\n",
    "\n",
    "# ุถุจุท ุฏูุฉ ุงูููุฏูู\n",
    "video_cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)\n",
    "video_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)\n",
    "\n",
    "print(\"๐ท ุงููุงููุฑุง ุชุนูู... ุงุถุบุท 'q' ููุฅุบูุงู\")\n",
    "\n",
    "# ๐ ุชุญููู ุฌููุน ุงูุตูุฑ ูู ูุฌูุฏ 'photo/' ูุงูุชุนุฑู ุนูู ุงููุฌูู\n",
    "face_encodings_list = []\n",
    "face_names = []\n",
    "\n",
    "photo_dir = \"face_id_photos/\"\n",
    "if not os.path.exists(photo_dir):\n",
    "    print(f\"โ๏ธ ุฎุทุฃ: ุงููุฌูุฏ '{photo_dir}' ุบูุฑ ููุฌูุฏ.\")\n",
    "    exit()\n",
    "\n",
    "for file_name in os.listdir(photo_dir):\n",
    "    if file_name.endswith((\".jpg\", \".jpeg\", \".png\")):\n",
    "        image_path = os.path.join(photo_dir, file_name)\n",
    "        image = face_recognition.load_image_file(image_path)\n",
    "        encodings = face_recognition.face_encodings(image)\n",
    "\n",
    "        if encodings:\n",
    "            face_encodings_list.append(encodings[0])\n",
    "            face_names.append(os.path.splitext(file_name)[0])\n",
    "        else:\n",
    "            print(f\"โ๏ธ ุชุญุฐูุฑ: ูู ูุชู ุงูุนุซูุฑ ุนูู ูุฌู ูู {file_name}\")\n",
    "\n",
    "if not face_encodings_list:\n",
    "    print(\"โ๏ธ ุฎุทุฃ: ูู ูุชู ุชุญููู ุฃู ูุฌููุ ุชุฃูุฏ ูู ูุถุน ุตูุฑ ุตุงูุญุฉ ูู ูุฌูุฏ 'photo/'.\")\n",
    "    exit()\n",
    "\n",
    "print(f\"ุชู ุชุญููู ุงูุฃุณูุงุก ุงูุชุงููุฉ: {face_names}\")\n",
    "\n",
    "# ูุชุญ ููู Excel ุงูููุฌูุฏ\n",
    "excel_file = 'Attendance.xlsx'\n",
    "if not os.path.exists(excel_file):\n",
    "    print(f\"โ๏ธ ุฎุทุฃ: ุงูููู {excel_file} ุบูุฑ ููุฌูุฏ.\")\n",
    "    exit()\n",
    "\n",
    "wb = openpyxl.load_workbook(excel_file)\n",
    "ws = wb.active\n",
    "\n",
    "# ุชุญุฏูุฏ ุนููุฏ ุงูุญุถูุฑ ููุฐุง ุงูููู\n",
    "current_date = datetime.now().strftime(\"%Y-%m-%d\")\n",
    "attendance_column = None\n",
    "\n",
    "# ุงูุจุญุซ ุนู ุนููุฏ ุจุชุงุฑูุฎ ุงูููู\n",
    "for col in range(2, ws.max_column + 1):  \n",
    "    if ws.cell(row=1, column=col).value == current_date:\n",
    "        attendance_column = col\n",
    "        break\n",
    "\n",
    "# ุฅุฐุง ูู ููู ููุฌูุฏูุงุ ูุชู ุฅูุดุงุก ุนููุฏ ุฌุฏูุฏ\n",
    "if not attendance_column:\n",
    "    attendance_column = ws.max_column + 1\n",
    "    ws.cell(row=1, column=attendance_column).value = current_date\n",
    "\n",
    "student_list = face_names.copy()\n",
    "\n",
    "try:\n",
    "    while True:\n",
    "        ret, frame = video_cap.read()\n",
    "        if not ret:\n",
    "            print(\"โ๏ธ ุฎุทุฃ: ูู ูุชู ุงูุชูุงุท ุตูุฑุฉ ูู ุงููุงููุฑุง.\")\n",
    "            break\n",
    "\n",
    "        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)\n",
    "        face_locations = face_recognition.face_locations(rgb)\n",
    "        print(f\"โ ุชู ุงูุชุดุงู {len(face_locations)} ูุฌู(ูุฌูู)\")\n",
    "\n",
    "        face_names_detected = []\n",
    "        if face_locations:\n",
    "            face_encodings = face_recognition.face_encodings(rgb, face_locations)\n",
    "            for face_encoding in face_encodings:\n",
    "                matches = face_recognition.compare_faces(face_encodings_list, face_encoding, tolerance=0.5)\n",
    "                face_distances = face_recognition.face_distance(face_encodings_list, face_encoding)\n",
    "                print(f\"๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: {face_distances}\")\n",
    "\n",
    "                if any(matches):  \n",
    "                    best_match = np.argmin(face_distances)\n",
    "                    name = face_names[best_match]\n",
    "                else:\n",
    "                    name = \"Unknown\"\n",
    "                \n",
    "                face_names_detected.append(name)\n",
    "                \n",
    "                if name in face_names and name in student_list:\n",
    "                    student_list.remove(name)\n",
    "                    \n",
    "                    # ุงูุจุญุซ ุนู ุงูุตู ุงูููุงุณุจ ููุงุณู\n",
    "                    for row in range(2, ws.max_row + 1):\n",
    "                        if ws.cell(row=row, column=1).value == name:\n",
    "                            # ุงูุชุญูู ููุง ุฅุฐุง ูุงู ุงูุญุถูุฑ ูุณุฌููุง ุจุงููุนู ูู ุนููุฏ ุงูููู\n",
    "                            if ws.cell(row=row, column=attendance_column).value is None:\n",
    "                                ws.cell(row=row, column=attendance_column).value = True\n",
    "                                print(f\"๐ ุชู ุชุณุฌูู ุงูุญุถูุฑ ููุทุงูุจ {name} ูู ููู Excel\")\n",
    "                            else:\n",
    "                                print(f\"โ {name} ูุณุฌู ุจุงููุนู ุงููููุ ูู ูุชู ุงูุชุบููุฑ\")\n",
    "                            break\n",
    "\n",
    "        for (top, right, bottom, left), name in zip(face_locations, face_names_detected):\n",
    "            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)\n",
    "            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)\n",
    "\n",
    "        cv2.imshow('Face Recognition', frame)\n",
    "\n",
    "        if cv2.waitKey(1) & 0xFF == ord('q'):\n",
    "            print(\"๐ด ุฅุบูุงู ุงููุงููุฑุง.\")\n",
    "            break\n",
    "except KeyboardInterrupt:\n",
    "    print(\"๐ด ุชู ุฅุบูุงู ุงูุจุฑูุงูุฌ.\")\n",
    "finally:\n",
    "    wb.save(excel_file)\n",
    "    video_cap.release()\n",
    "    cv2.destroyAllWindows()\n",
    "    print(\"โ ุชู ุฅุบูุงู ุงููุงููุฑุง ูุฅููุงู ุงูุจุฑูุงูุฌ ุจูุฌุงุญ.\")\n"
   ]
  },
  {
   "cell_type": "markdown",
   "id": "efeaf362",
   "metadata": {},
   "source": [
    "# Importing Required Libraries"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 2,
   "id": "f800b026",
   "metadata": {},
   "outputs": [],
   "source": [
    "import cv2\n",
    "\n",
    "cap = cv2.VideoCapture(1)  # ุงุณุชุฎุฏู 0 ูููุงููุฑุง ุงูุงูุชุฑุงุถูุฉุ ุฃู 1 ููุงููุฑุง ุฃุฎุฑู\n",
    "\n",
    "while True:\n",
    "    ret, frame = cap.read()\n",
    "    if not ret:\n",
    "        break\n",
    "\n",
    "    cv2.imshow(\"Camera Feed\", frame)\n",
    "\n",
    "    if cv2.waitKey(1) & 0xFF == ord('q'):  # ุงุถุบุท 'q' ููุฎุฑูุฌ\n",
    "        break\n",
    "\n",
    "cap.release()\n",
    "cv2.destroyAllWindows()\n",
    "\n"
   ]
  },
  {
   "cell_type": "markdown",
   "id": "23e462a4",
   "metadata": {},
   "source": [
    "# Opening the Webcam and Checking \n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 3,
   "id": "508f9aa3",
   "metadata": {},
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "Camera found at index 0\n",
      "Total connected cameras: 1\n"
     ]
    },
    {
     "data": {
      "text/plain": [
       "1"
      ]
     },
     "execution_count": 3,
     "metadata": {},
     "output_type": "execute_result"
    }
   ],
   "source": [
    "import cv2\n",
    "\n",
    "def count_connected_cameras():\n",
    "    camera_count = 0\n",
    "    for i in range(10):  # ุงูุญุต ุญุชู 10 ููุงูุฐ (ููููู ุฒูุงุฏุชูุง ุญุณุจ ุงูุญุงุฌุฉ)\n",
    "        cap = cv2.VideoCapture(i)\n",
    "        if cap.isOpened():\n",
    "            print(f\"Camera found at index {i}\")\n",
    "            camera_count += 1\n",
    "            cap.release()\n",
    "    \n",
    "    print(f\"Total connected cameras: {camera_count}\")\n",
    "    return camera_count\n",
    "\n",
    "count_connected_cameras()\n"
   ]
  },
  {
   "cell_type": "markdown",
   "id": "1b280c6f",
   "metadata": {},
   "source": [
    "# Setting the Camera Resolution And  Loading Stored Faces for Recognition"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 9,
   "id": "3dd586fb",
   "metadata": {},
   "outputs": [],
   "source": []
  },
  {
   "cell_type": "markdown",
   "id": "3e1c5ac1",
   "metadata": {},
   "source": [
    "# Identifying the Attendance Column for Today"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 10,
   "id": "64a66164",
   "metadata": {},
   "outputs": [],
   "source": [
    "\n"
   ]
  },
  {
   "cell_type": "markdown",
   "id": "4df1e390",
   "metadata": {},
   "source": [
    "# Capturing Video and Detecting Faces And Marking Attendance in the Excel File And  Drawing Boxes Around Detected Faces"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "03f29ef2",
   "metadata": {},
   "outputs": [],
   "source": [
    "\n"
   ]
  },
  {
   "cell_type": "markdown",
   "id": "5905023f",
   "metadata": {},
   "source": [
    "# Closing the Camera and Saving Data"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 2,
   "id": "df646d6f",
   "metadata": {},
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "๐ท ุงููุงููุฑุง ุชุนูู... ุงุถุบุท 'q' ููุฅุบูุงู\n",
      "ุชู ุชุญููู ุงูุฃุณูุงุก ุงูุชุงููุฉ: ['eman', 'kareem', 'mahoud', 'mohamed', 'mostafa', 'nada', 'youisf']\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.789184   0.38827478 0.61576336 0.64127109 0.65972947 0.74285169\n",
      " 0.6415747 ]\n",
      "๐ ุชู ุชุณุฌูู ุงูุญุถูุฑ ููุทุงูุจ kareem ูู ููู Excel\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.79401931 0.39067163 0.62237242 0.63972602 0.67292461 0.75715364\n",
      " 0.63224882]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.79124651 0.35611407 0.61809494 0.64021339 0.69969764 0.72795133\n",
      " 0.63411332]\n",
      "โ ุชู ุงูุชุดุงู 1 ูุฌู(ูุฌูู)\n",
      "๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: [0.81003007 0.35041894 0.61567755 0.62771409 0.70358528 0.73970752\n",
      " 0.65227404]\n",
      "๐ด ุฅุบูุงู ุงููุงููุฑุง.\n",
      "๐ด ุงูุทุงูุจ eman ูู ูุญุถุฑ ุฏุฑุณ ุงูููู.\n",
      "๐ด ุงูุทุงูุจ mahoud ูู ูุญุถุฑ ุฏุฑุณ ุงูููู.\n",
      "๐ด ุงูุทุงูุจ mohamed ูู ูุญุถุฑ ุฏุฑุณ ุงูููู.\n",
      "๐ด ุงูุทุงูุจ mostafa ูู ูุญุถุฑ ุฏุฑุณ ุงูููู.\n",
      "๐ด ุงูุทุงูุจ nada ูู ูุญุถุฑ ุฏุฑุณ ุงูููู.\n",
      "๐ด ุงูุทุงูุจ youisf ูู ูุญุถุฑ ุฏุฑุณ ุงูููู.\n",
      "๐ ุงูุทุงูุจ nada ูู ูููู ุญุถูุฑ ุงูุฏุฑูุณ ุงููุทููุจ ูุฐุง ุงูุดูุฑ (1/6).\n",
      "๐ ุงูุทุงูุจ youisf ูู ูููู ุญุถูุฑ ุงูุฏุฑูุณ ุงููุทููุจ ูุฐุง ุงูุดูุฑ (0/6).\n",
      "โ ุชู ุฅุบูุงู ุงููุงููุฑุง ูุฅููุงู ุงูุจุฑูุงูุฌ ุจูุฌุงุญ.\n"
     ]
    }
   ],
   "source": [
    "import cv2\n",
    "import face_recognition\n",
    "import openpyxl\n",
    "from datetime import datetime\n",
    "import numpy as np\n",
    "import os\n",
    "\n",
    "# ุจุฏุก ุงูุชูุงุท ุงูููุฏูู ูู ุงููุงููุฑุง\n",
    "video_cap = cv2.VideoCapture(0)\n",
    "\n",
    "# ุงูุชุญูู ูู ุฃู ุงููุงููุฑุง ุชุนูู\n",
    "if not video_cap.isOpened():\n",
    "    print(\"โ๏ธ ุฎุทุฃ: ุชุนุฐุฑ ูุชุญ ุงููุงููุฑุง.\")\n",
    "    exit()\n",
    "\n",
    "# ุถุจุท ุฏูุฉ ุงูููุฏูู\n",
    "video_cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)\n",
    "video_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)\n",
    "\n",
    "print(\"๐ท ุงููุงููุฑุง ุชุนูู... ุงุถุบุท 'q' ููุฅุบูุงู\")\n",
    "\n",
    "# ๐ ุชุญููู ุฌููุน ุงูุตูุฑ ูู ูุฌูุฏ 'face_id_photos/' ูุงูุชุนุฑู ุนูู ุงููุฌูู\n",
    "face_encodings_list = []\n",
    "face_names = []\n",
    "\n",
    "photo_dir = \"photo/\"\n",
    "if not os.path.exists(photo_dir):\n",
    "    print(f\"โ๏ธ ุฎุทุฃ: ุงููุฌูุฏ '{photo_dir}' ุบูุฑ ููุฌูุฏ.\")\n",
    "    exit()\n",
    "\n",
    "for file_name in os.listdir(photo_dir):\n",
    "    if file_name.endswith((\".jpg\", \".jpeg\", \".png\")):\n",
    "        image_path = os.path.join(photo_dir, file_name)\n",
    "        image = face_recognition.load_image_file(image_path)\n",
    "        encodings = face_recognition.face_encodings(image)\n",
    "\n",
    "        if encodings:\n",
    "            face_encodings_list.append(encodings[0])\n",
    "            face_names.append(os.path.splitext(file_name)[0])\n",
    "        else:\n",
    "            print(f\"โ๏ธ ุชุญุฐูุฑ: ูู ูุชู ุงูุนุซูุฑ ุนูู ูุฌู ูู {file_name}\")\n",
    "\n",
    "if not face_encodings_list:\n",
    "    print(\"โ๏ธ ุฎุทุฃ: ูู ูุชู ุชุญููู ุฃู ูุฌููุ ุชุฃูุฏ ูู ูุถุน ุตูุฑ ุตุงูุญุฉ ูู ุงููุฌูุฏ.\")\n",
    "    exit()\n",
    "\n",
    "print(f\"ุชู ุชุญููู ุงูุฃุณูุงุก ุงูุชุงููุฉ: {face_names}\")\n",
    "\n",
    "# ูุชุญ ููู Excel ุงูููุฌูุฏ\n",
    "excel_file = 'Attendance.xlsx'\n",
    "if not os.path.exists(excel_file):\n",
    "    print(f\"โ๏ธ ุฎุทุฃ: ุงูููู {excel_file} ุบูุฑ ููุฌูุฏ.\")\n",
    "    exit()\n",
    "\n",
    "wb = openpyxl.load_workbook(excel_file)\n",
    "ws = wb.active\n",
    "\n",
    "# ุชุญุฏูุฏ ุนููุฏ ุงูุญุถูุฑ ููุฐุง ุงูููู\n",
    "current_date = datetime.now().strftime(\"%Y-%m-%d\")\n",
    "attendance_column = None\n",
    "\n",
    "# ุงูุจุญุซ ุนู ุนููุฏ ุจุชุงุฑูุฎ ุงูููู\n",
    "for col in range(2, ws.max_column + 1):\n",
    "    if ws.cell(row=1, column=col).value == current_date:\n",
    "        attendance_column = col\n",
    "        break\n",
    "\n",
    "# ุฅุฐุง ูู ููู ููุฌูุฏูุงุ ูุชู ุฅูุดุงุก ุนููุฏ ุฌุฏูุฏ\n",
    "if not attendance_column:\n",
    "    attendance_column = ws.max_column + 1\n",
    "    ws.cell(row=1, column=attendance_column).value = current_date\n",
    "\n",
    "# ูุงุฆูุฉ ุงูุทูุงุจ ุงูููุชุฑุถ ุชุณุฌูู ุญุถูุฑูู (ูุณุชูุฏุฉ ูู ุฃุณูุงุก ุงููุฌูู)\n",
    "student_list = face_names.copy()\n",
    "\n",
    "try:\n",
    "    while True:\n",
    "        ret, frame = video_cap.read()\n",
    "        if not ret:\n",
    "            print(\"โ๏ธ ุฎุทุฃ: ูู ูุชู ุงูุชูุงุท ุตูุฑุฉ ูู ุงููุงููุฑุง.\")\n",
    "            break\n",
    "\n",
    "        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)\n",
    "        face_locations = face_recognition.face_locations(rgb)\n",
    "        print(f\"โ ุชู ุงูุชุดุงู {len(face_locations)} ูุฌู(ูุฌูู)\")\n",
    "\n",
    "        face_names_detected = []\n",
    "        if face_locations:\n",
    "            face_encodings = face_recognition.face_encodings(rgb, face_locations)\n",
    "            for face_encoding in face_encodings:\n",
    "                matches = face_recognition.compare_faces(face_encodings_list, face_encoding, tolerance=0.5)\n",
    "                face_distances = face_recognition.face_distance(face_encodings_list, face_encoding)\n",
    "                print(f\"๐ ุงููุณุงูุงุช ุจูู ุงููุฌูู: {face_distances}\")\n",
    "\n",
    "                if any(matches):\n",
    "                    best_match = np.argmin(face_distances)\n",
    "                    name = face_names[best_match]\n",
    "                else:\n",
    "                    name = \"Unknown\"\n",
    "                \n",
    "                face_names_detected.append(name)\n",
    "                \n",
    "                # ุชุณุฌูู ุงูุญุถูุฑ ูู ููู Excel ุฅุฐุง ูุงู ุงูุงุณู ูุนุฑูููุง ููู ูุชู ุชุณุฌููู ูุณุจููุง\n",
    "                if name in face_names and name in student_list:\n",
    "                    student_list.remove(name)\n",
    "                    \n",
    "                    # ุงูุจุญุซ ุนู ุงูุตู ุงูููุงุณุจ ููุงุณู ูู ููู Excel\n",
    "                    for row in range(2, ws.max_row + 1):\n",
    "                        if ws.cell(row=row, column=1).value == name:\n",
    "                            # ุงูุชุญูู ููุง ุฅุฐุง ูุงู ุงูุญุถูุฑ ูุณุฌููุง ุจุงููุนู ูู ุนููุฏ ุงูููู\n",
    "                            if ws.cell(row=row, column=attendance_column).value is None:\n",
    "                                ws.cell(row=row, column=attendance_column).value = True\n",
    "                                print(f\"๐ ุชู ุชุณุฌูู ุงูุญุถูุฑ ููุทุงูุจ {name} ูู ููู Excel\")\n",
    "                            else:\n",
    "                                print(f\"โ {name} ูุณุฌู ุจุงููุนู ุงููููุ ูู ูุชู ุงูุชุบููุฑ\")\n",
    "                            break\n",
    "\n",
    "        # ุฑุณู ูุณุชุทูู ุญูู ุงููุฌูู ููุชุงุจุฉ ุงูุงุณู\n",
    "        for (top, right, bottom, left), name in zip(face_locations, face_names_detected):\n",
    "            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)\n",
    "            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)\n",
    "\n",
    "        cv2.imshow('Face Recognition', frame)\n",
    "\n",
    "        if cv2.waitKey(1) & 0xFF == ord('q'):\n",
    "            print(\"๐ด ุฅุบูุงู ุงููุงููุฑุง.\")\n",
    "            break\n",
    "except KeyboardInterrupt:\n",
    "    print(\"๐ด ุชู ุฅุบูุงู ุงูุจุฑูุงูุฌ.\")\n",
    "finally:\n",
    "    # ุงูุชุญูู ูู ุงูุทูุงุจ ุงูุฐูู ูู ูุญุถุฑูุง ุฏุฑุณ ุงูููู\n",
    "    if student_list:\n",
    "        for absent_student in student_list:\n",
    "            print(f\"๐ด ุงูุทุงูุจ {absent_student} ูู ูุญุถุฑ ุฏุฑุณ ุงูููู.\")\n",
    "            # ููููู ููุง ุฅุถุงูุฉ ุงูููุฏ ูุฅุฑุณุงู ุฑุณุงูุฉ ุชูุจูู ููุทุงูุจ (ูุซูุงู ุนุจุฑ ุงูุจุฑูุฏ ุงูุฅููุชุฑููู)\n",
    "\n",
    "    # ุงูุชุญูู ูู ุนุฏุฏ ุงูุญุถูุฑ ุฎูุงู ุงูุดูุฑ ููู ุทุงูุจ\n",
    "    for row in range(2, ws.max_row + 1):\n",
    "        student_name = ws.cell(row=row, column=1).value\n",
    "        attendance_count = 0\n",
    "        # ููุชุฑุถ ุฃู ุงูุฃุนูุฏุฉ ูู 2 ุญุชู ุขุฎุฑ ุนููุฏ ุชูุซู ุงูุฏุฑูุณ ูู ูุฐุง ุงูุดูุฑ\n",
    "        for col in range(2, ws.max_column + 1):\n",
    "            if ws.cell(row=row, column=col).value == True:\n",
    "                attendance_count += 1\n",
    "        if attendance_count < 6:\n",
    "            print(f\"๐ ุงูุทุงูุจ {student_name} ูู ูููู ุญุถูุฑ ุงูุฏุฑูุณ ุงููุทููุจ ูุฐุง ุงูุดูุฑ ({attendance_count}/6).\")\n",
    "            # ููููู ููุง ุฅุถุงูุฉ ุงูููุฏ ูุฅุฑุณุงู ุฑุณุงูุฉ ุชูุจูู ููุทุงูุจ\n",
    "\n",
    "    wb.save(excel_file)\n",
    "    video_cap.release()\n",
    "    cv2.destroyAllWindows()\n",
    "    print(\"โ ุชู ุฅุบูุงู ุงููุงููุฑุง ูุฅููุงู ุงูุจุฑูุงูุฌ ุจูุฌุงุญ.\")\n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "fb34ec06",
   "metadata": {},
   "outputs": [],
   "source": []
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "e6a08f03",
   "metadata": {},
   "outputs": [],
   "source": []
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "c2e027ec",
   "metadata": {},
   "outputs": [],
   "source": []
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "483d4555",
   "metadata": {},
   "outputs": [],
   "source": []
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python 3 (ipykernel)",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.9.13"
  }
 },
 "nbformat": 4,
 "nbformat_minor": 5
}
